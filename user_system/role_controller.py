# user_system/role_controller.py

from flask import request, jsonify, g, Blueprint, send_file  # Añadido send_file
from db_config import get_connection
from utils.session_validator import session_validator
from utils.auditoria import registrar_auditoria
# importaciones para la descarga de pdf y excel

import pdfkit
from flask import render_template, request, send_file

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from datetime import datetime
import io

roles_bp = Blueprint('roles', __name__)


@roles_bp.route('/roles/consulta', methods=['GET'])
@session_validator(tabla="usuarios", accion="read")
def obtener_roles():
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            # Consulta simplificada
            cursor.execute("""
                SELECT 
                    r.idRol, 
                    r.nombreRol, 
                    r.descripcion,
                    (SELECT COUNT(*) FROM rol_permisos rp WHERE rp.idRol = r.idRol) AS totalPermisos
                FROM roles r
            """)

            roles = []
            for row in cursor:
                roles.append({
                    "idRol": row[0],
                    "nombreRol": row[1],
                    "descripcion": row[2] if row[2] is not None else "NO DESCRIPCION",
                    "totalPermisos": row[3]
                })

            # Depuración
            print("Roles mapeados manualmente:")
            for rol in roles:
                print(rol)

            # Para cada rol, obtener permisos
            for rol in roles:
                with conn.cursor() as perm_cursor:
                    perm_cursor.execute("""
                        SELECT p.tabla, GROUP_CONCAT(p.accion) AS acciones
                        FROM rol_permisos rp
                        JOIN permisos p ON rp.idPermiso = p.idPermiso
                        WHERE rp.idRol = %s
                        GROUP BY p.tabla
                        LIMIT 3
                    """, (rol['idRol'],))

                    permisos = {}
                    for (tabla, acciones) in perm_cursor:
                        permisos[tabla] = acciones.split(',')
                    rol['permisos'] = permisos

            return jsonify(roles), 200
    except Exception as e:
        print("Error al obtener roles:", e)
        return jsonify({"error": "Error al obtener roles"}), 500
    finally:
        conn.close()


@roles_bp.route('/roles', methods=['POST'])
@session_validator(tabla="usuarios", accion="create")
def crear_rol():
    data = request.json
    if 'nombreRol' not in data:
        return jsonify({'error': 'Falta el campo nombreRol'}), 400

    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                "INSERT INTO roles (nombreRol, descripcion) VALUES (%s, %s)",
                (data['nombreRol'], data.get('descripcion', ''))
            )
            nuevo_id = cursor.lastrowid
            conn.commit()

            # Obtener el rol recién creado
            cursor.execute("""
                SELECT idRol, nombreRol, descripcion 
                FROM roles 
                WHERE idRol = %s
            """, (nuevo_id,))
            nuevo_rol = dict(zip(
                [desc[0] for desc in cursor.description],
                cursor.fetchone()
            ))

            # Auditoría
            registrar_auditoria(
                g.user_id,
                'create',
                'roles',
                nuevo_id,
                valores_anteriores=None,
                valores_nuevos=data
            )

            return jsonify({
                'mensaje': 'Rol creado',
                'rol': nuevo_rol
            }), 201
    except Exception as e:
        conn.rollback()
        print(f"Error al crear rol: {str(e)}")
        return jsonify({'error': 'Error al crear rol'}), 500
    finally:
        conn.close()


@roles_bp.route('/roles/<int:id_rol>', methods=['PUT'])
@session_validator(tabla="usuarios", accion="update")
def actualizar_rol(id_rol):
    data = request.json
    if 'nombreRol' not in data:
        return jsonify({'error': 'Falta el campo nombreRol'}), 400

    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            # Obtener rol actual
            cursor.execute("""
                SELECT nombreRol, descripcion 
                FROM roles 
                WHERE idRol = %s
            """, (id_rol,))
            resultado = cursor.fetchone()

            if not resultado:
                return jsonify({'error': 'Rol no encontrado'}), 404

            column_names = [desc[0] for desc in cursor.description]
            rol_actual = dict(zip(column_names, resultado))

            # Actualizar rol
            cursor.execute("""
                UPDATE roles 
                SET nombreRol = %s, descripcion = %s 
                WHERE idRol = %s
            """, (data['nombreRol'], data.get('descripcion', ''), id_rol))
            conn.commit()

            # Obtener rol actualizado
            cursor.execute("""
                SELECT idRol, nombreRol, descripcion 
                FROM roles 
                WHERE idRol = %s
            """, (id_rol,))
            rol_actualizado = dict(zip(
                [desc[0] for desc in cursor.description],
                cursor.fetchone()
            ))

            # Preparar datos para auditoría
            valores_anteriores = {
                'nombreRol': rol_actual['nombreRol'],
                'descripcion': rol_actual['descripcion']
            }

            valores_nuevos = {
                'nombreRol': data['nombreRol'],
                'descripcion': data.get('descripcion', '')
            }

            # Auditoría
            registrar_auditoria(
                g.user_id,
                'update',
                'roles',
                id_rol,
                valores_anteriores=valores_anteriores,
                valores_nuevos=valores_nuevos
            )

            return jsonify({
                'mensaje': 'Rol actualizado',
                'rol': rol_actualizado
            }), 200
    except Exception as e:
        conn.rollback()
        print(f"Error al actualizar rol: {str(e)}")
        return jsonify({'error': 'Error al actualizar rol'}), 500
    finally:
        conn.close()


@roles_bp.route('/roles/<int:id_rol>', methods=['DELETE'])
@session_validator(tabla="usuarios", accion="delete")
def eliminar_rol(id_rol):
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            # Verificar si el rol está en uso
            cursor.execute("SELECT COUNT(*) AS count FROM usuarios WHERE idRol = %s", (id_rol,))
            resultado = cursor.fetchone()
            count = resultado[0] if resultado else 0

            if count > 0:
                return jsonify({
                    'error': 'No se puede eliminar, el rol está asignado a usuarios'
                }), 400

            # Obtener rol completo
            cursor.execute("""
                SELECT idRol, nombreRol, descripcion 
                FROM roles 
                WHERE idRol = %s
            """, (id_rol,))
            resultado = cursor.fetchone()

            if not resultado:
                return jsonify({'error': 'Rol no encontrado'}), 404

            column_names = [desc[0] for desc in cursor.description]
            rol = dict(zip(column_names, resultado))

            # Eliminar
            cursor.execute("DELETE FROM roles WHERE idRol = %s", (id_rol,))
            conn.commit()

            # Auditoría
            registrar_auditoria(
                g.user_id,
                'delete',
                'roles',
                id_rol,
                valores_anteriores=rol,
                valores_nuevos=None
            )

            return jsonify({
                'mensaje': 'Rol eliminado',
                'rol': rol
            }), 200
    except Exception as e:
        conn.rollback()
        print(f"Error al eliminar rol: {str(e)}")
        return jsonify({'error': 'Error al eliminar rol'}), 500
    finally:
        conn.close()


@roles_bp.route('/roles/<int:id_rol>/permisos', methods=['GET'])
@session_validator(tabla="permisos", accion="read")
def obtener_permisos_rol(id_rol):
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            # Verificar existencia del rol
            cursor.execute("SELECT idRol FROM roles WHERE idRol = %s", (id_rol,))
            if not cursor.fetchone():
                return jsonify({'error': 'Rol no encontrado'}), 404

            # Obtener permisos asignados
            cursor.execute("""
                SELECT 
                    p.idPermiso,
                    p.tabla,
                    p.accion
                FROM rol_permisos rp
                JOIN permisos p ON rp.idPermiso = p.idPermiso
                WHERE rp.idRol = %s
            """, (id_rol,))

            permisos = []
            column_names = [desc[0] for desc in cursor.description]
            for row in cursor.fetchall():
                permisos.append(dict(zip(column_names, row)))

            return jsonify(permisos), 200
    except Exception as e:
        print(f"Error al obtener permisos del rol: {str(e)}")
        return jsonify({'error': 'Error al obtener permisos del rol'}), 500
    finally:
        conn.close()

# roles asignados a usuarios
@roles_bp.route('/roles/<int:id_rol>/usuarios', methods=['GET'])
@session_validator(tabla="usuarios", accion="read")
def obtener_usuarios_por_rol(id_rol):
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            # Verificar existencia del rol
            cursor.execute("SELECT idRol FROM roles WHERE idRol = %s", (id_rol,))
            if not cursor.fetchone():
                return jsonify({'error': 'Rol no encontrado'}), 404

            # Obtener usuarios con sus datos completos
            cursor.execute("""
                SELECT 
                    u.idUsuario,
                    u.nombre,
                    u.apellidop,
                    u.apellidom,
                    u.email,
                    u.telefono,
                    u.estatus,
                    u.foto
                FROM usuarios u
                WHERE u.idRol = %s
            """, (id_rol,))

            usuarios = []
            column_names = [desc[0] for desc in cursor.description]
            for row in cursor.fetchall():
                usuario = dict(zip(column_names, row))
                usuarios.append(usuario)

            return jsonify(usuarios), 200
    except Exception as e:
        print(f"Error al obtener usuarios del rol: {str(e)}")
        return jsonify({'error': 'Error al obtener usuarios del rol'}), 500
    finally:
        conn.close()


@roles_bp.route('/permisos/disponibles', methods=['GET'])
@session_validator(tabla="permisos", accion="read")
def obtener_permisos_disponibles():
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    idPermiso,
                    tabla,
                    accion
                FROM permisos
                ORDER BY tabla, accion
            """)

            permisos = []
            column_names = [desc[0] for desc in cursor.description]
            for row in cursor.fetchall():
                permisos.append(dict(zip(column_names, row)))

            # Organizar por tabla
            permisos_por_tabla = {}
            for permiso in permisos:
                tabla = permiso['tabla']
                if tabla not in permisos_por_tabla:
                    permisos_por_tabla[tabla] = []
                permisos_por_tabla[tabla].append(permiso['accion'])

            return jsonify(permisos_por_tabla), 200
    except Exception as e:
        print(f"Error al obtener permisos disponibles: {str(e)}")
        return jsonify({'error': 'Error al obtener permisos disponibles'}), 500
    finally:
        conn.close()


@roles_bp.route('/permisos', methods=['POST'])
@session_validator(tabla="permisos", accion="update")
def gestionar_permisos():
    data = request.json
    required_fields = ['tipoDestino', 'idDestino', 'tabla', 'accion', 'operacion']

    # Validar campos requeridos
    if not all(field in data for field in required_fields):
        return jsonify({'error': 'Faltan campos requeridos'}), 400

    # Validar valores permitidos
    if data['tipoDestino'] not in ['usuario', 'rol']:
        return jsonify({'error': 'tipoDestino debe ser "usuario" o "rol"'}), 400

    if data['operacion'] not in ['asignar', 'revocar']:
        return jsonify({'error': 'operacion debe ser "asignar" o "revocar"'}), 400

    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            # Llamar al stored procedure
            cursor.callproc('sp_GestionarPermiso', (
                data['tipoDestino'],
                data['idDestino'],
                data['tabla'],
                data['accion'],
                data['operacion']
            ))
            conn.commit()

            return jsonify({'mensaje': 'Permiso gestionado correctamente'}), 200
    except Exception as e:
        conn.rollback()
        error_msg = str(e)
        print(f"Error al gestionar permiso: {error_msg}")

        # Manejar errores específicos del stored procedure
        if 'Rol no encontrado' in error_msg:
            return jsonify({'error': 'Rol no encontrado'}), 404
        elif 'Usuario no encontrado' in error_msg:
            return jsonify({'error': 'Usuario no encontrado'}), 404
        elif 'El permiso especificado no existe' in error_msg:
            return jsonify({'error': 'Permiso no encontrado'}), 404
        else:
            return jsonify({'error': 'Error al gestionar permiso'}), 500
    finally:
        conn.close()


# Función auxiliar para obtener todos los roles con permisos
def obtener_todos_los_roles():
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    r.idRol, 
                    r.nombreRol, 
                    r.descripcion,
                    (SELECT COUNT(*) FROM rol_permisos rp WHERE rp.idRol = r.idRol) AS totalPermisos
                FROM roles r
            """)
            roles = []
            for row in cursor:
                roles.append({
                    "idRol": row[0],
                    "nombreRol": row[1],
                    "descripcion": row[2] if row[2] is not None else "NO DESCRIPCION",
                    "totalPermisos": row[3]
                })

            # Obtener permisos para cada rol
            for rol in roles:
                with conn.cursor() as perm_cursor:
                    perm_cursor.execute("""
                        SELECT p.tabla, GROUP_CONCAT(p.accion) AS acciones
                        FROM rol_permisos rp
                        JOIN permisos p ON rp.idPermiso = p.idPermiso
                        WHERE rp.idRol = %s
                        GROUP BY p.tabla
                    """, (rol['idRol'],))

                    permisos = {}
                    for (tabla, acciones) in perm_cursor:
                        permisos[tabla] = acciones.split(',')
                    rol['permisos'] = permisos

            return roles
    except Exception as e:
        print("Error al obtener roles:", e)
        return []
    finally:
        conn.close()

# Exportar a PDF (solo nombre y descripción)
@roles_bp.route('/roles/exportar/pdf', methods=['GET'])
@session_validator(tabla="roles", accion="read")
def exportar_roles_pdf():
    try:
        conn = get_connection()
        roles = []
        try:
            with conn.cursor() as cursor:
                cursor.execute("""
                    SELECT nombreRol, descripcion
                    FROM roles
                """)
                for row in cursor:
                    roles.append({
                        "nombreRol": row[0],
                        "descripcion": row[1] if row[1] else "Sin descripción"
                    })
        finally:
            conn.close()

        html = render_template(
            "roles_reporte.html",  # Ajusta la ruta si es diferente
            roles=roles,
            fecha=datetime.now().strftime("%d/%m/%Y %H:%M"),
            titulo="Reporte de Roles"
        )

        config = pdfkit.configuration(wkhtmltopdf='/usr/local/bin/wkhtmltopdf')

        pdf = pdfkit.from_string(html, False, configuration=config)

        return send_file(
            io.BytesIO(pdf),
            as_attachment=True,
            download_name=f'reporte_roles_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf',
            mimetype='application/pdf'
        )

    except Exception as e:
        import traceback
        print("Error generando PDF:", e)
        traceback.print_exc()
        return jsonify({'error': 'Error generando reporte PDF: ' + str(e)}), 500


# Exportar a Excel (solo nombre y descripción)
@roles_bp.route('/roles/exportar/excel', methods=['GET'])
@session_validator(tabla="roles", accion="read")
def exportar_roles_excel():
    try:
        conn = get_connection()
        try:
            with conn.cursor() as cursor:
                # Obtener solo nombre y descripción
                cursor.execute("""
                    SELECT nombreRol, descripcion
                    FROM roles
                """)
                roles = []
                for row in cursor:
                    roles.append({
                        "nombreRol": row[0],
                        "descripcion": row[1] if row[1] is not None else "Sin descripción"
                    })
        finally:
            conn.close()

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Roles"

        # Estilos
        header_fill = PatternFill(start_color='3b82f6', end_color='3b82f6', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        alignment = Alignment(horizontal='center', vertical='center')
        border = Border(bottom=Side(style='medium'))

        # Encabezados (solo 2 columnas)
        headers = [
            "Rol",
            "Descripción"
        ]

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f'{col_letter}1']
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = border

        # Datos
        for row_num, role in enumerate(roles, 2):
            ws.cell(row=row_num, column=1, value=role['nombreRol'])
            ws.cell(row=row_num, column=2, value=role['descripcion'])

        # Ajustar anchos
        column_widths = [30, 70]  # Solo para las 2 columnas
        for i, width in enumerate(column_widths, 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = width

        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'reporte_roles_{datetime.datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Error generando Excel: {str(e)}")
        return jsonify({'error': 'Error generando reporte Excel: ' + str(e)}), 500