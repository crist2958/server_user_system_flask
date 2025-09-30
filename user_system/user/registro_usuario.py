# /user_system/user/registro_usuario.py

from flask import jsonify, g, Blueprint
import bcrypt
import os
from werkzeug.utils import secure_filename
from db_config import get_connection
from utils.session_validator import session_validator
from utils.auditoria import registrar_auditoria

# importaciones para la descarga de pdf y excel

import pdfkit
from flask import render_template, request, send_file

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


from datetime import datetime
import io


# Crear blueprints
registro_bp = Blueprint('registro', __name__)
usuarios_bp = Blueprint('usuarios', __name__)
procedures_bp = Blueprint('procedures', __name__)
archivos_bp = Blueprint('archivo', __name__)


# Función auxiliar para verificar superusuarios
def es_superusuario(id_usuario):
    """Verifica si un usuario es superusuario"""
    conexion = get_connection()
    try:
        with conexion.cursor() as cursor:
            cursor.execute("SELECT is_superadmin FROM usuarios WHERE idUsuario = %s", (id_usuario,))
            result = cursor.fetchone()
            return result and result[0]  # True si es superusuario
    except Exception as e:
        print(f"Error al verificar superusuario: {e}")
        return False
    finally:
        conexion.close()


# Funciones para manejar archivos
def subir_archivo(tabla, id_registro, archivo, campo, carpeta, user_id_actor=None):
    # Crear directorio si no existe
    ruta_directorio = os.path.join('uploads', carpeta)
    os.makedirs(ruta_directorio, exist_ok=True)

    # Eliminar archivo anterior si existe
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(f"SELECT {campo} FROM {tabla} WHERE id{tabla[:-1].capitalize()} = %s", (id_registro,))
            archivo_anterior = cursor.fetchone()

            if archivo_anterior and archivo_anterior[0]:
                eliminar_archivo(archivo_anterior[0], carpeta)
    finally:
        conn.close()

    # Subir nuevo archivo
    nombre_seguro = secure_filename(archivo.filename)
    nombre_archivo = f'{id_registro}_{nombre_seguro}'

    ruta_completa = os.path.join(ruta_directorio, nombre_archivo)
    archivo.save(ruta_completa)

    # Actualizar base de datos
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(f"""
                UPDATE {tabla}
                SET {campo} = %s
                WHERE id{tabla[:-1].capitalize()} = %s
            """, (nombre_archivo, id_registro))
            conn.commit()
    finally:
        conn.close()

    return nombre_archivo


def eliminar_archivo(nombre_archivo, carpeta):
    if not nombre_archivo:
        return False

    try:
        ruta_archivo = os.path.join('uploads', carpeta, nombre_archivo)
        if os.path.exists(ruta_archivo):
            os.remove(ruta_archivo)
            return True
        return False
    except Exception as e:
        print(f"Error al eliminar archivo: {e}")
        return False


# Endpoint para registrar usuario
@registro_bp.route('/usuarios/registro', methods=['POST'])
@session_validator(tabla="usuarios", accion="create")
def registrar_usuario():
    datos = request.json
    campos_requeridos = [
        'nombreUsuario', 'nombre', 'apellidop',
        'apellidom', 'email', 'password', 'idRol'
    ]

    # 1) Validar que todos los campos requeridos estén presentes
    if not all(c in datos for c in campos_requeridos):
        return jsonify({'error': 'Faltan campos requeridos'}), 400

    # 2) Hashear la contraseña
    password_bytes = datos['password'].encode('utf-8')
    salt = bcrypt.gensalt()
    password_hash = bcrypt.hashpw(password_bytes, salt).decode('utf-8')

    # Obtener teléfono si está presente (campo opcional)
    telefono = datos.get('telefono', None)  # Usar .get() para evitar KeyError

    conexion = get_connection()
    try:
        with conexion.cursor() as cursor:
            # 3) INSERT incluyendo idRol y teléfono
            cursor.execute("""
                INSERT INTO usuarios 
                    (nombreUsuario, nombre, apellidop, apellidom, 
                     email, telefono, password_hash, idRol)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                datos['nombreUsuario'],
                datos['nombre'],
                datos['apellidop'],
                datos['apellidom'],
                datos['email'],
                telefono,  # Campo de teléfono añadido
                password_hash,
                datos['idRol']
            ))
            conexion.commit()

            nuevo_id = cursor.lastrowid

            # 4) Registrar auditoría con todos los valores
            id_usuario_actor = getattr(g, 'user_id', None)
            registrar_auditoria(
                id_usuario_actor,
                'create',
                'usuarios',
                nuevo_id,
                valores_anteriores=None,
                valores_nuevos={
                    'nombreUsuario': datos['nombreUsuario'],
                    'nombre': datos['nombre'],
                    'apellidop': datos['apellidop'],
                    'apellidom': datos['apellidom'],
                    'email': datos['email'],
                    'telefono': telefono,  # Incluir teléfono en auditoría
                    'idRol': datos['idRol']
                }
            )

        return jsonify({'mensaje': 'Usuario registrado correctamente', 'id': nuevo_id}), 201

    except Exception as e:
        conexion.rollback()
        print("Error al registrar usuario:", e)
        return jsonify({'error': 'Error al registrar usuario: ' + str(e)}), 500

    finally:
        conexion.close()


# Endpoint para actualizar usuario
@registro_bp.route('/usuarios/<int:id_usuario>', methods=['PUT'])
@session_validator(tabla="usuarios", accion="update")
def actualizar_usuario(id_usuario):
    # Protección para superusuarios
    if es_superusuario(id_usuario):
        return jsonify({'error': 'No se pueden modificar superusuarios'}), 403

    datos = request.json
    campos_requeridos = ['nombreUsuario', 'nombre', 'apellidop', 'apellidom', 'email', 'idRol']

    if not all(c in datos for c in campos_requeridos):
        return jsonify({'error': 'Faltan campos requeridos'}), 400

    # Obtener teléfono (campo opcional)
    telefono = datos.get('telefono', None)

    # Manejo de contraseña (opcional)
    password = datos.get('password', None)
    password_hash = None
    if password:
        password_bytes = password.encode('utf-8')
        salt = bcrypt.gensalt()
        password_hash = bcrypt.hashpw(password_bytes, salt).decode('utf-8')

    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Obtener valores anteriores
            cursor.execute("SELECT * FROM usuarios WHERE idUsuario = %s", (id_usuario,))
            usuario_anterior = cursor.fetchone()
            if not usuario_anterior:
                return jsonify({'error': 'Usuario no encontrado'}), 404

            # Preparar datos para auditoría
            valores_anteriores = {
                'nombreUsuario': usuario_anterior['nombreUsuario'],
                'nombre': usuario_anterior['nombre'],
                'apellidop': usuario_anterior['apellidop'],
                'apellidom': usuario_anterior['apellidom'],
                'email': usuario_anterior['email'],
                'telefono': usuario_anterior['telefono'],
                'idRol': usuario_anterior['idRol']
            }

            valores_nuevos = {
                'nombreUsuario': datos['nombreUsuario'],
                'nombre': datos['nombre'],
                'apellidop': datos['apellidop'],
                'apellidom': datos['apellidom'],
                'email': datos['email'],
                'telefono': telefono,
                'idRol': datos['idRol']
            }

            # Construir consulta SQL dinámica
            sql = """
                UPDATE usuarios
                SET nombreUsuario = %s, 
                    nombre = %s, 
                    apellidop = %s, 
                    apellidom = %s,
                    email = %s, 
                    telefono = %s,
                    idRol = %s
            """
            params = [
                datos['nombreUsuario'],
                datos['nombre'],
                datos['apellidop'],
                datos['apellidom'],
                datos['email'],
                telefono,
                datos['idRol'],
                id_usuario
            ]

            # Agregar contraseña si fue proporcionada
            if password_hash:
                sql = sql + ", password_hash = %s"
                params.insert(-1, password_hash)
                # Registrar cambio de contraseña en auditoría
                valores_anteriores['password'] = '******'
                valores_nuevos['password'] = '******'

            sql = sql + " WHERE idUsuario = %s"

            # Ejecutar actualización
            cursor.execute(sql, params)
            conexion.commit()

            # Auditoría
            id_usuario_actor = getattr(g, 'user_id', None)
            registrar_auditoria(
                id_usuario_actor,
                'update',
                'usuarios',
                id_usuario,
                valores_anteriores=valores_anteriores,
                valores_nuevos=valores_nuevos
            )

        return jsonify({'mensaje': 'Usuario actualizado correctamente'}), 200

    except Exception as e:
        conexion.rollback()
        print("Error al actualizar usuario:", e)
        return jsonify({'error': 'Error al actualizar usuario: ' + str(e)}), 500
    finally:
        conexion.close()


# Endpoint para eliminar usuario
@registro_bp.route('/usuarios/<int:id_usuario>', methods=['DELETE'])
@session_validator(tabla="usuarios", accion="delete")
def eliminar_usuario(id_usuario):
    # Protección para superusuarios
    if es_superusuario(id_usuario):
        return jsonify({'error': 'No se pueden eliminar superusuarios'}), 403

    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Obtener datos antes de eliminar
            cursor.execute("SELECT * FROM usuarios WHERE idUsuario = %s", (id_usuario,))
            usuario = cursor.fetchone()
            if not usuario:
                return jsonify({'error': 'Usuario no encontrado'}), 404

            # Eliminar usuarios
            cursor.execute("DELETE FROM usuarios WHERE idUsuario = %s", (id_usuario,))
            conexion.commit()

            # Auditoría
            id_usuario_actor = getattr(g, 'user_id', None)
            registrar_auditoria(
                id_usuario_actor,
                'delete',
                'usuarios',
                id_usuario,
                valores_anteriores={
                    k: usuario[k] for k in usuario if k != 'password_hash'
                },
                valores_nuevos=None
            )

        return jsonify({'mensaje': 'Usuario eliminado correctamente'}), 200

    except Exception as e:
        conexion.rollback()
        print("Error al eliminar usuarios:", e)
        return jsonify({'error': 'Error al eliminar usuarios'}), 500

    finally:
        conexion.close()


# Endpoint para subir foto de perfil
@registro_bp.route('/usuarios/<int:id_usuario>/foto', methods=['POST'])
@session_validator(tabla="usuarios", accion="update")
def subir_foto_usuario(id_usuario):
    # Protección para superusuarios
    if es_superusuario(id_usuario):
        return jsonify({'error': 'No se pueden modificar fotos de superusuarios'}), 403

    if 'imagen' not in request.files:
        return jsonify({'error': 'No se envió ninguna imagen'}), 400

    archivo = request.files['imagen']

    if archivo.filename == '':
        return jsonify({'error': 'Nombre de archivo vacío'}), 400

    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Obtener foto anterior
            cursor.execute("SELECT foto FROM usuarios WHERE idUsuario = %s", (id_usuario,))
            usuario = cursor.fetchone()

            if not usuario:
                return jsonify({'error': 'Usuario no encontrado'}), 404

            foto_anterior = usuario.get('foto')

        # Subir nuevo archivo
        nombre_archivo = subir_archivo(
            tabla='usuarios',
            id_registro=id_usuario,
            archivo=archivo,
            campo='foto',
            carpeta='usuarios',
            user_id_actor=g.user_id
        )

        # Registrar auditoría
        registrar_auditoria(
            g.user_id,
            'update',
            'usuarios',
            id_usuario,
            valores_anteriores={'foto': foto_anterior},
            valores_nuevos={'foto': nombre_archivo}
        )

        return jsonify({
            'mensaje': 'Foto de perfil actualizada correctamente',
            'nombre_archivo': nombre_archivo,
            'ruta': f"/archivo/usuarios/{id_usuario}/foto"
        }), 200

    except Exception as e:
        conexion.rollback()
        print("Error al subir la foto de perfil:", e)
        return jsonify({'error': 'Error al subir la foto de perfil'}), 500

    finally:
        conexion.close()


# Endpoint para eliminar foto de perfil
@registro_bp.route('/usuarios/<int:id_usuario>/foto', methods=['DELETE'])
@session_validator(tabla="usuarios", accion="update")
def eliminar_foto_usuario(id_usuario):
    # Protección para superusuarios
    if es_superusuario(id_usuario):
        return jsonify({'error': 'No se pueden eliminar fotos de superusuarios'}), 403

    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Obtener foto actual
            cursor.execute("SELECT foto FROM usuarios WHERE idUsuario = %s", (id_usuario,))
            usuario = cursor.fetchone()

            if not usuario:
                return jsonify({'error': 'Usuario no encontrado'}), 404

            foto_actual = usuario.get('foto')

            if not foto_actual:
                return jsonify({'mensaje': 'El usuario no tiene foto'}), 200

            # Eliminar archivo físico
            if eliminar_archivo(foto_actual, 'usuarios'):
                # Actualizar base de datos
                cursor.execute(
                    "UPDATE usuarios SET foto = NULL WHERE idUsuario = %s",
                    (id_usuario,)
                )
                conexion.commit()

                # Auditoría
                registrar_auditoria(
                    g.user_id,
                    'update',
                    'usuarios',
                    id_usuario,
                    valores_anteriores={'foto': foto_actual},
                    valores_nuevos={'foto': None}
                )

                return jsonify({'mensaje': 'Foto eliminada correctamente'}), 200
            else:
                return jsonify({'error': 'No se pudo eliminar el archivo'}), 500

    except Exception as e:
        conexion.rollback()
        print("Error al eliminar foto:", e)
        return jsonify({'error': 'Error al eliminar la foto'}), 500
    finally:
        conexion.close()


# Endpoint para obtener usuarios
@usuarios_bp.route('/usuarios', methods=['GET'])
@session_validator(tabla="usuarios", accion="read")
def obtener_usuarios():
    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Excluir superusuarios del listado
            cursor.execute("SELECT * FROM vw_usuarios_con_roles WHERE is_superadmin = 0")
            usuarios = cursor.fetchall()

            for u in usuarios:
                if u['foto']:
                    u['foto_url'] = f"/archivo/usuarios/{u['idUsuario']}/foto"
                else:
                    u['foto_url'] = None

        return jsonify(usuarios), 200

    except Exception as e:
        print("Error al obtener usuarios:", e)
        return jsonify({"error": "Error al obtener usuarios"}), 500

    finally:
        conexion.close()


# Endpoint para obtener roles
@usuarios_bp.route('/roles', methods=['GET'])
@session_validator(tabla="usuarios", accion="read")
def obtener_roles():
    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            cursor.execute("SELECT idRol, nombreRol FROM roles")
            roles = cursor.fetchall()
        return jsonify(roles), 200
    except Exception as e:
        print("Error al obtener roles:", e)
        return jsonify({"error": "Error al obtener roles"}), 500
    finally:
        conexion.close()


# Endpoint para cambiar estado de usuario
@usuarios_bp.route('/usuarios/<int:id_usuario>/estado', methods=['PATCH'])
@session_validator(tabla="usuarios", accion="update")
def cambiar_estado_usuario(id_usuario):
    # Protección para superusuarios
    if es_superusuario(id_usuario):
        return jsonify({'error': 'No se puede desactivar superusuarios'}), 403

    data = request.get_json()
    if 'estatus' not in data:
        return jsonify({'error': 'Falta el campo estatus'}), 400

    nuevo_estatus = data['estatus']
    if nuevo_estatus not in ['Activo', 'Inactivo']:
        return jsonify({'error': 'Estatus no válido'}), 400

    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Obtener el estado anterior
            cursor.execute("SELECT estatus FROM usuarios WHERE idUsuario = %s", (id_usuario,))
            usuario = cursor.fetchone()
            if not usuario:
                return jsonify({'error': 'Usuario no encontrado'}), 404

            # Actualizar
            cursor.execute("UPDATE usuarios SET estatus = %s WHERE idUsuario = %s", (nuevo_estatus, id_usuario))
            conexion.commit()

            # Auditoría
            id_usuario_actor = getattr(g, 'user_id', None)
            registrar_auditoria(
                id_usuario_actor,
                'update',
                'usuarios',
                id_usuario,
                valores_anteriores={'estatus': usuario['estatus']},
                valores_nuevos={'estatus': nuevo_estatus}
            )

        return jsonify({'mensaje': 'Estado actualizado'}), 200

    except Exception as e:
        conexion.rollback()
        print("Error al actualizar estado:", e)
        return jsonify({'error': 'Error al actualizar estado'}), 500

    finally:
        conexion.close()


# Endpoint para obtener permisos disponibles
@usuarios_bp.route('/permisos/disponibles', methods=['GET'])
@session_validator(tabla="permisos", accion="read")
def obtener_permisos_disponibles():
    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            cursor.execute("SELECT DISTINCT tabla FROM permisos")
            tablas = [row['tabla'] for row in cursor.fetchall()]

            # Obtener acciones para cada tabla
            permisos_disponibles = {}
            for tabla in tablas:
                cursor.execute("""
                    SELECT DISTINCT accion 
                    FROM permisos 
                    WHERE tabla = %s
                """, (tabla,))
                acciones = [row['accion'] for row in cursor.fetchall()]
                permisos_disponibles[tabla] = acciones

            return jsonify(permisos_disponibles), 200
    except Exception as e:
        print("Error al obtener permisos disponibles:", e)
        return jsonify({"error": "Error al obtener permisos disponibles"}), 500
    finally:
        conexion.close()


# Endpoint para obtener permisos directos de usuario
@usuarios_bp.route('/usuarios/<int:id_usuario>/permisos-directos', methods=['GET'])
@session_validator(tabla="permisos", accion="read")
def obtener_permisos_directos_usuario(id_usuario):
    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            cursor.execute("""
                SELECT p.tabla, p.accion 
                FROM usuario_permisos up
                JOIN permisos p ON up.idPermiso = p.idPermiso
                WHERE up.idUsuario = %s
            """, (id_usuario,))

            permisos = {}
            for row in cursor.fetchall():
                if row['tabla'] not in permisos:
                    permisos[row['tabla']] = []
                permisos[row['tabla']].append(row['accion'])

            return jsonify(permisos), 200
    except Exception as e:
        print("Error al obtener permisos directos:", e)
        return jsonify({"error": "Error al obtener permisos directos"}), 500
    finally:
        conexion.close()


# Endpoint para actualizar permisos
@procedures_bp.route('/permisos', methods=['POST'])
@session_validator(tabla="permisos", accion="update")
def ejecutar_procedimiento_permisos():
    data = request.json
    required_params = ['tipoDestino', 'idDestino', 'tabla', 'accion', 'operacion']

    # Validar parámetros requeridos
    if not all(param in data for param in required_params):
        return jsonify({'error': 'Faltan parámetros requeridos'}), 400

    # Validar valores de enumeraciones
    if data['tipoDestino'] not in ['usuario', 'rol']:
        return jsonify({'error': 'tipoDestino debe ser "usuario" o "rol"'}), 400

    if data['operacion'] not in ['asignar', 'revocar']:
        return jsonify({'error': 'operacion debe ser "asignar" o "revocar"'}), 400

    conexion = get_connection()
    try:
        with conexion.cursor() as cursor:
            # Ejecutar el procedimiento almacenado
            cursor.callproc('sp_GestionarPermiso', [
                data['tipoDestino'],
                data['idDestino'],
                data['tabla'],
                data['accion'],
                data['operacion']
            ])
            conexion.commit()

            # Registrar auditoría
            registrar_auditoria(
                g.user_id,
                'update',
                'permisos',
                None,  # No hay ID específico
                valores_anteriores=None,
                valores_nuevos=data
            )

            return jsonify({'mensaje': 'Operación completada exitosamente'}), 200

    except Exception as e:
        conexion.rollback()
        print(f"Error al ejecutar procedimiento: {str(e)}")
        return jsonify({'error': f'Error en el procedimiento: {str(e)}'}), 500
    finally:
        conexion.close()


# Endpoint para servir fotos de perfil
@archivos_bp.route('/archivo/usuarios/<int:id_usuario>/foto', methods=['GET'])
def servir_foto_usuario(id_usuario):
    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            cursor.execute("SELECT foto FROM usuarios WHERE idUsuario = %s", (id_usuario,))
            usuario = cursor.fetchone()

            if not usuario or not usuario['foto']:
                # Puedes devolver una imagen por defecto aquí si lo prefieres
                return jsonify({'error': 'Foto no encontrada'}), 404

            nombre_archivo = usuario['foto']
            ruta_archivo = os.path.join('uploads', 'usuarios', nombre_archivo)

            if not os.path.exists(ruta_archivo):
                return jsonify({'error': 'Archivo no encontrado'}), 404

            return send_file(ruta_archivo, mimetype='image/jpeg')

    except Exception as e:
        print("Error al servir archivo:", e)
        return jsonify({"error": "Error al obtener foto"}), 500
    finally:
        conexion.close()

'''
# Descargar PDF
@usuarios_bp.route('/usuarios/exportar/pdf', methods=['GET'])
@session_validator(tabla="usuarios", accion="read")
def exportar_usuarios_pdf():
    try:
        # Obtener parámetros de filtrado
        search = request.args.get('search', '')
        status_filter = request.args.get('status', 'all')
        sort_by = request.args.get('sort', 'name')

        # Obtener usuarios (con misma lógica que en el front)
        usuarios = obtener_usuarios_filtrados(search, status_filter, sort_by)

        # Crear PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            title="Reporte de Usuarios",
            author="Sistema de Gestión"
        )

        styles = getSampleStyleSheet()
        elements = []

        # Encabezado
        elements.append(Paragraph("REPORTE DE USUARIOS", styles['Title']))
        elements.append(Paragraph(f"Fecha: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 12))

        # Tabla de datos
        headers = [
            "Usuario",
            "Nombre Completo",
            "Correo",
            "Teléfono",
            "Fecha Registro",
            "Estado",
            "Rol"
        ]

        data = [headers]
        for usuario in usuarios:
            nombre_completo = f"{usuario['nombre']} {usuario['apellidop']} {usuario['apellidom']}"
            estado = "Activo" if usuario['estatus'] == "Activo" else "Inactivo"

            # Manejar fechas nulas
            fecha_registro = usuario['fechaRegistro']
            fecha_str = fecha_registro.strftime('%d/%m/%Y') if fecha_registro else "N/A"

            data.append([
                usuario['nombreUsuario'],
                nombre_completo,
                usuario['email'],
                usuario['telefono'] or "N/A",
                fecha_str,
                estado,
                usuario['rol']  # Corregido: campo correcto es 'rol'
            ])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'reporte_usuarios_{datetime.datetime.now().strftime("%Y%m%d_%H%M")}.pdf',
            mimetype='application/pdf'
        )

    except Exception as e:
        print(f"Error generando PDF: {str(e)}")
        return jsonify({'error': 'Error generando reporte PDF: ' + str(e)}), 500
'''




@usuarios_bp.route('/usuarios/exportar/pdf', methods=['GET'])
@session_validator(tabla="usuarios", accion="read")
def exportar_usuarios_pdf():
    try:
        # Parámetros de filtrado
        search = request.args.get('search', '')
        status = request.args.get('status', 'all')
        sort = request.args.get('sort', 'name')

        usuarios = obtener_usuarios_filtrados(search, status, sort)

        html = render_template(
            "usuarios_reporte.html",
            usuarios=usuarios,
            fecha=datetime.now().strftime("%d/%m/%Y %H:%M"),
            titulo="Reporte de Usuarios"
        )

        # Ruta explícita de wkhtmltopdf (importante en macOS)
        config = pdfkit.configuration(wkhtmltopdf='/usr/local/bin/wkhtmltopdf')

        pdf = pdfkit.from_string(html, False, configuration=config)

        return send_file(
            io.BytesIO(pdf),
            as_attachment=True,
            download_name=f'reporte_usuarios_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf',
            mimetype='application/pdf'
        )

    except Exception as e:
        import traceback
        print("Error al generar el PDF:", e)
        traceback.print_exc()
        return jsonify({'error': 'Error generando reporte PDF: ' + str(e)}), 500


# Descargar excel
@usuarios_bp.route('/usuarios/exportar/excel', methods=['GET'])
@session_validator(tabla="usuarios", accion="read")
def exportar_usuarios_excel():
    try:
        # Obtener parámetros de filtrado
        search = request.args.get('search', '')
        status_filter = request.args.get('status', 'all')
        sort_by = request.args.get('sort', 'name')

        # Obtener usuarios
        usuarios = obtener_usuarios_filtrados(search, status_filter, sort_by)

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Usuarios"

        # Estilos
        header_fill = PatternFill(start_color='3b82f6', end_color='3b82f6', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        alignment = Alignment(horizontal='center', vertical='center')
        border = Border(bottom=Side(style='medium'))

        # Encabezados
        headers = [
            "Usuario",
            "Nombre",
            "Apellido Paterno",
            "Apellido Materno",
            "Correo",
            "Teléfono",
            "Fecha Registro",
            "Estado",
            "Rol"
        ]

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f'{col_letter}1']
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = border

        # Datos
        for row_num, usuario in enumerate(usuarios, 2):
            # Manejar fechas nulas
            fecha_registro = usuario['fechaRegistro']
            fecha_str = fecha_registro.strftime('%d/%m/%Y') if fecha_registro else "N/A"

            ws.cell(row=row_num, column=1, value=usuario['nombreUsuario'])
            ws.cell(row=row_num, column=2, value=usuario['nombre'])
            ws.cell(row=row_num, column=3, value=usuario['apellidop'])
            ws.cell(row=row_num, column=4, value=usuario['apellidom'])
            ws.cell(row=row_num, column=5, value=usuario['email'])
            ws.cell(row=row_num, column=6, value=usuario['telefono'] or "")
            ws.cell(row=row_num, column=7, value=fecha_str)
            ws.cell(row=row_num, column=8, value="Activo" if usuario['estatus'] == "Activo" else "Inactivo")
            ws.cell(row=row_num, column=9, value=usuario['rol'])  # Corregido: campo correcto es 'rol'

        # Ajustar anchos
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 20

        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'reporte_usuarios_{datetime.datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Error generando Excel: {str(e)}")
        return jsonify({'error': 'Error generando reporte Excel: ' + str(e)}), 500


def obtener_usuarios_filtrados(search, status_filter, sort_by):
    """Función auxiliar para obtener usuarios con filtros"""
    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            query = """
                SELECT * 
                FROM vw_usuarios_con_roles
                WHERE is_superadmin = 0  -- Excluir superusuarios
            """
            params = []

            # Filtro de búsqueda
            if search:
                query += """
                    AND (nombreUsuario LIKE %s 
                    OR nombre LIKE %s 
                    OR apellidop LIKE %s 
                    OR apellidom LIKE %s 
                    OR email LIKE %s 
                    OR rol LIKE %s)  /* Corregido: campo correcto es rol */
                """
                search_term = f"%{search}%"
                params.extend([search_term] * 6)

            # Filtro de estado
            if status_filter == 'active':
                query += " AND estatus = 'Activo'"
            elif status_filter == 'inactive':
                query += " AND estatus = 'Inactivo'"

            # Ordenamiento
            if sort_by == 'name':
                query += " ORDER BY nombre"
            elif sort_by == 'date':
                query += " ORDER BY fechaRegistro DESC"
            elif sort_by == 'role':
                query += " ORDER BY rol"  # Corregido: campo correcto es rol

            cursor.execute(query, params)
            return cursor.fetchall()

    except Exception as e:
        print(f"Error obteniendo usuarios: {str(e)}")
        return []
    finally:
        conexion.close()
