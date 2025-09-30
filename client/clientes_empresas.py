# client/clientes_empresas.py

from flask import jsonify, g, Blueprint
from db_config import get_connection
from utils.session_validator import session_validator
from utils.auditoria import registrar_auditoria
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

import pdfkit
from flask import render_template, request, send_file

empresas_bp = Blueprint('empresas', __name__)

# Endpoint para registrar empresa
@empresas_bp.route('/empresas', methods=['POST'])
@session_validator(tabla="clientes", accion="create")
def registrar_empresa():
    datos = request.json
    campos_requeridos = ['nombre', 'rfc', 'telefono', 'email', 'razonSocial', 'domicilioFiscal']

    if not all(campo in datos for campo in campos_requeridos):
        return jsonify({'error': 'Faltan campos requeridos'}), 400

    conexion = get_connection()
    try:
        with conexion.cursor() as cursor:
            # Paso 1: Insertar en tabla clientes
            cursor.execute("""
                INSERT INTO clientes (
                    tipoCliente, nombre, rfc, telefono, 
                    email, direccion, notas
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                'Empresa',
                datos['nombre'],
                datos['rfc'],
                datos['telefono'],
                datos['email'],
                datos.get('direccion'),
                datos.get('notas')
            ))
            id_cliente = cursor.lastrowid

            # Registrar auditoría para cliente
            registrar_auditoria(
                g.user_id,
                'create',
                'clientes',
                id_cliente,
                valores_anteriores=None,
                valores_nuevos=datos
            )

            # Paso 2: Insertar en tabla empresas
            cursor.execute("""
                INSERT INTO empresas (
                    idCliente, razonSocial, telefonoContacto, 
                    domicilioFiscal, numFacturacion
                ) VALUES (%s, %s, %s, %s, %s)
            """, (
                id_cliente,
                datos['razonSocial'],
                datos.get('telefonoContacto', datos['telefono']),
                datos['domicilioFiscal'],
                datos.get('numFacturacion')
            ))
            id_empresa = cursor.lastrowid

            # Registrar auditoría para empresa
            registrar_auditoria(
                g.user_id,
                'create',
                'empresas',
                id_empresa,
                valores_anteriores=None,
                valores_nuevos={
                    'idCliente': id_cliente,
                    'razonSocial': datos['razonSocial'],
                    'domicilioFiscal': datos['domicilioFiscal']
                }
            )

            # Paso 3: Insertar contactos si existen
            contactos = datos.get('contactos', [])
            for contacto in contactos:
                cursor.execute("""
                    INSERT INTO contacto (
                        idEmpresa, nombre, apellidoP, apellidoM, 
                        Area, telefono, email
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (
                    id_empresa,
                    contacto['nombre'],
                    contacto.get('apellidoP'),
                    contacto.get('apellidoM'),
                    contacto.get('Area'),
                    contacto['telefono'],
                    contacto['email']
                ))
                id_contacto = cursor.lastrowid

                # Registrar auditoría para contacto
                registrar_auditoria(
                    g.user_id,
                    'create',
                    'contacto',
                    id_contacto,
                    valores_anteriores=None,
                    valores_nuevos=contacto
                )

            # Paso 4: Insertar direcciones
            direcciones = datos.get('direcciones', [])
            for direccion in direcciones:
                cursor.execute("""
                    INSERT INTO direcciones_envio (
                        idCliente, calle, numero, colonia, 
                        codigoPostal, pais, estado, ciudad
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    id_cliente,
                    direccion['calle'],
                    direccion.get('numero'),
                    direccion.get('colonia'),
                    direccion['codigoPostal'],
                    direccion['pais'],
                    direccion['estado'],
                    direccion['ciudad']
                ))

            conexion.commit()
            return jsonify({
                'mensaje': 'Empresa registrada correctamente',
                'idCliente': id_cliente,
                'idEmpresa': id_empresa
            }), 201

    except Exception as e:
        conexion.rollback()
        print("Error al registrar empresa:", e)
        return jsonify({'error': 'Error al registrar empresa: ' + str(e)}), 500
    finally:
        conexion.close()

# Endpoint para actualizar empresa
@empresas_bp.route('/empresas/<int:id_cliente>', methods=['PUT'])
@session_validator(tabla="clientes", accion="update")
def actualizar_empresa(id_cliente):
    datos = request.json
    conexion = get_connection()

    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Verificar que sea una empresa
            cursor.execute("SELECT tipoCliente FROM clientes WHERE idCliente = %s", (id_cliente,))
            cliente = cursor.fetchone()

            if not cliente:
                return jsonify({'error': 'Cliente no encontrado'}), 404
            if cliente['tipoCliente'] != 'Empresa':
                return jsonify({'error': 'El cliente no es una empresa'}), 400

            # Obtener empresa
            cursor.execute("SELECT * FROM empresas WHERE idCliente = %s", (id_cliente,))
            empresa = cursor.fetchone()
            if not empresa:
                return jsonify({'error': 'Empresa no encontrada'}), 404

            # Actualizar datos del cliente
            campos_cliente = [
                'nombre', 'rfc', 'telefono', 'email',
                'direccion', 'notas', 'estatus'
            ]

            valores_anteriores = {}
            valores_nuevos = {}
            set_clause = []
            params = []

            for campo in campos_cliente:
                if campo in datos:
                    cursor.execute(f"SELECT {campo} FROM clientes WHERE idCliente = %s", (id_cliente,))
                    valor_anterior = cursor.fetchone()[campo]

                    valores_anteriores[campo] = valor_anterior
                    valores_nuevos[campo] = datos[campo]
                    set_clause.append(f"{campo} = %s")
                    params.append(datos[campo])

            if set_clause:
                sql = f"UPDATE clientes SET {', '.join(set_clause)} WHERE idCliente = %s"
                params.append(id_cliente)
                cursor.execute(sql, params)

                # Registrar auditoría para cliente
                registrar_auditoria(
                    g.user_id,
                    'update',
                    'clientes',
                    id_cliente,
                    valores_anteriores=valores_anteriores,
                    valores_nuevos=valores_nuevos
                )

            # Actualizar datos de la empresa
            campos_empresa = [
                'razonSocial', 'telefonoContacto',
                'domicilioFiscal', 'numFacturacion'
            ]

            valores_anteriores_emp = {}
            valores_nuevos_emp = {}
            set_clause_emp = []
            params_emp = []

            for campo in campos_empresa:
                if campo in datos:
                    cursor.execute(f"SELECT {campo} FROM empresas WHERE idEmpresa = %s", (empresa['idEmpresa'],))
                    valor_anterior = cursor.fetchone()[campo]

                    valores_anteriores_emp[campo] = valor_anterior
                    valores_nuevos_emp[campo] = datos[campo]
                    set_clause_emp.append(f"{campo} = %s")
                    params_emp.append(datos[campo])

            if set_clause_emp:
                sql_emp = f"UPDATE empresas SET {', '.join(set_clause_emp)} WHERE idEmpresa = %s"
                params_emp.append(empresa['idEmpresa'])
                cursor.execute(sql_emp, params_emp)

                # Registrar auditoría para empresa
                registrar_auditoria(
                    g.user_id,
                    'update',
                    'empresas',
                    empresa['idEmpresa'],
                    valores_anteriores=valores_anteriores_emp,
                    valores_nuevos=valores_nuevos_emp
                )

            conexion.commit()
            return jsonify({'mensaje': 'Empresa actualizada correctamente'}), 200

    except Exception as e:
        conexion.rollback()
        print("Error al actualizar empresa:", e)
        return jsonify({'error': 'Error al actualizar empresa: ' + str(e)}), 500
    finally:
        conexion.close()

# Endpoint para eliminar empresa
@empresas_bp.route('/empresas/<int:id_cliente>', methods=['DELETE'])
@session_validator(tabla="clientes", accion="delete")
def eliminar_empresa(id_cliente):
    conexion = get_connection()

    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Verificar que sea empresa
            cursor.execute("SELECT * FROM clientes WHERE idCliente = %s", (id_cliente,))
            cliente = cursor.fetchone()

            if not cliente:
                return jsonify({'error': 'Cliente no encontrado'}), 404
            if cliente['tipoCliente'] != 'Empresa':
                return jsonify({'error': 'El cliente no es una empresa'}), 400

            # Eliminar cliente (CASCADE eliminará empresas, contactos y direcciones)
            cursor.execute("DELETE FROM clientes WHERE idCliente = %s", (id_cliente,))
            conexion.commit()

            # Registrar auditoría
            registrar_auditoria(
                g.user_id,
                'delete',
                'clientes',
                id_cliente,
                valores_anteriores=cliente,
                valores_nuevos=None
            )

            return jsonify({'mensaje': 'Empresa eliminada correctamente'}), 200

    except Exception as e:
        conexion.rollback()
        print("Error al eliminar empresa:", e)
        return jsonify({'error': 'Error al eliminar empresa: ' + str(e)}), 500
    finally:
        conexion.close()

# Endpoint para listar empresas (ACTUALIZADO: incluye idEmpresa)
@empresas_bp.route('/empresas', methods=['GET'])
@session_validator(tabla="clientes", accion="read")
def listar_empresas():
    search = request.args.get('search', '')
    estatus = request.args.get('estatus', 'all')

    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            query = """
                SELECT c.*, e.idEmpresa, e.razonSocial, e.domicilioFiscal,
                       COUNT(d.idDireccionEnvio) AS num_direcciones,
                       COUNT(cont.idContacto) AS num_contactos
                FROM clientes c
                JOIN empresas e ON c.idCliente = e.idCliente
                LEFT JOIN direcciones_envio d ON c.idCliente = d.idCliente
                LEFT JOIN contacto cont ON e.idEmpresa = cont.idEmpresa
                WHERE c.tipoCliente = 'Empresa'
            """
            params = []

            # Filtro de búsqueda
            if search:
                query += " AND (c.nombre LIKE %s OR c.rfc LIKE %s OR c.email LIKE %s OR e.razonSocial LIKE %s)"
                search_term = f"%{search}%"
                params.extend([search_term] * 4)

            # Filtro de estatus
            if estatus != 'all':
                query += " AND c.estatus = %s"
                params.append(estatus)

            query += " GROUP BY c.idCliente ORDER BY c.fechaRegistro DESC"

            cursor.execute(query, params)
            empresas = cursor.fetchall()

            # Formatear resultados
            for empresa in empresas:
                empresa['num_direcciones'] = int(empresa['num_direcciones'])
                empresa['num_contactos'] = int(empresa['num_contactos'])

            return jsonify(empresas), 200

    except Exception as e:
        print("Error al listar empresas:", e)
        return jsonify({"error": "Error al listar empresas"}), 500
    finally:
        conexion.close()

# Endpoint para obtener detalles de empresa (ACTUALIZADO)
@empresas_bp.route('/empresas/<int:id_cliente>', methods=['GET'])
@session_validator(tabla="clientes", accion="read")
def obtener_empresa(id_cliente):
    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Obtener datos básicos
            cursor.execute("""
                SELECT c.*, e.* 
                FROM clientes c
                JOIN empresas e ON c.idCliente = e.idCliente
                WHERE c.idCliente = %s AND c.tipoCliente = 'Empresa'
            """, (id_cliente,))
            empresa = cursor.fetchone()

            if not empresa:
                return jsonify({'error': 'Empresa no encontrada'}), 404

            # Obtener contactos usando idEmpresa
            cursor.execute("""
                SELECT * FROM contacto 
                WHERE idEmpresa = %s
            """, (empresa['idEmpresa'],))
            contactos = cursor.fetchall()

            # Obtener direcciones
            cursor.execute("""
                SELECT * FROM direcciones_envio 
                WHERE idCliente = %s
            """, (id_cliente,))
            direcciones = cursor.fetchall()

            return jsonify({
                'empresa': empresa,
                'contactos': contactos,
                'direcciones': direcciones
            }), 200

    except Exception as e:
        print("Error al obtener empresa:", e)
        return jsonify({"error": "Error al obtener empresa"}), 500
    finally:
        conexion.close()

# Endpoint para agregar contacto a empresa (ACTUALIZADO: verifica existencia de empresa)
@empresas_bp.route('/empresas/<int:id_empresa>/contactos', methods=['POST'])
@session_validator(tabla="clientes", accion="create")
def agregar_contacto_empresa(id_empresa):
    datos = request.json
    campos_requeridos = ['nombre', 'telefono', 'email']

    if not all(campo in datos for campo in campos_requeridos):
        return jsonify({'error': 'Faltan campos requeridos'}), 400

    conexion = get_connection()
    try:
        with conexion.cursor() as cursor:
            # Verificar existencia de la empresa
            cursor.execute("SELECT idEmpresa FROM empresas WHERE idEmpresa = %s", (id_empresa,))
            if not cursor.fetchone():
                return jsonify({'error': 'La empresa especificada no existe'}), 404

            # Insertar contacto
            cursor.execute("""
                INSERT INTO contacto (
                    idEmpresa, nombre, apellidoP, apellidoM, 
                    Area, telefono, email
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                id_empresa,
                datos['nombre'],
                datos.get('apellidoP'),
                datos.get('apellidoM'),
                datos.get('Area'),
                datos['telefono'],
                datos['email']
            ))
            id_contacto = cursor.lastrowid
            conexion.commit()

            # Registrar auditoría
            registrar_auditoria(
                g.user_id,
                'create',
                'contacto',
                id_contacto,
                valores_anteriores=None,
                valores_nuevos=datos
            )

            return jsonify({
                'mensaje': 'Contacto agregado correctamente',
                'idContacto': id_contacto
            }), 201

    except Exception as e:
        conexion.rollback()
        print("Error al agregar contacto:", e)
        return jsonify({'error': 'Error al agregar contacto: ' + str(e)}), 500
    finally:
        conexion.close()

# Endpoint para obtener contactos de una empresa específica (NUEVO)
@empresas_bp.route('/empresas/<int:id_empresa>/contactos', methods=['GET'])
@session_validator(tabla="clientes", accion="read")
def obtener_contactos_empresa(id_empresa):
    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Verificar existencia de la empresa
            cursor.execute("SELECT idEmpresa FROM empresas WHERE idEmpresa = %s", (id_empresa,))
            if not cursor.fetchone():
                return jsonify({'error': 'La empresa especificada no existe'}), 404

            # Obtener contactos
            cursor.execute("SELECT * FROM contacto WHERE idEmpresa = %s", (id_empresa,))
            contactos = cursor.fetchall()
            return jsonify(contactos), 200

    except Exception as e:
        print("Error al obtener contactos:", e)
        return jsonify({"error": "Error al obtener contactos"}), 500
    finally:
        conexion.close()

# Endpoint para actualizar contacto
@empresas_bp.route('/contactos/<int:id_contacto>', methods=['PUT'])
@session_validator(tabla="clientes", accion="update")
def actualizar_contacto(id_contacto):
    datos = request.json
    conexion = get_connection()

    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Obtener contacto actual
            cursor.execute("SELECT * FROM contacto WHERE idContacto = %s", (id_contacto,))
            contacto = cursor.fetchone()

            if not contacto:
                return jsonify({'error': 'Contacto no encontrado'}), 404

            # Actualizar campos
            campos_actualizables = [
                'nombre', 'apellidoP', 'apellidoM', 'Area',
                'telefono', 'email'
            ]

            valores_anteriores = {}
            valores_nuevos = {}
            set_clause = []
            params = []

            for campo in campos_actualizables:
                if campo in datos:
                    valores_anteriores[campo] = contacto.get(campo)
                    valores_nuevos[campo] = datos[campo]
                    set_clause.append(f"{campo} = %s")
                    params.append(datos[campo])

            if set_clause:
                sql = f"UPDATE contacto SET {', '.join(set_clause)} WHERE idContacto = %s"
                params.append(id_contacto)
                cursor.execute(sql, params)

                # Registrar auditoría
                registrar_auditoria(
                    g.user_id,
                    'update',
                    'contacto',
                    id_contacto,
                    valores_anteriores=valores_anteriores,
                    valores_nuevos=valores_nuevos
                )

            conexion.commit()
            return jsonify({'mensaje': 'Contacto actualizado correctamente'}), 200

    except Exception as e:
        conexion.rollback()
        print("Error al actualizar contacto:", e)
        return jsonify({'error': 'Error al actualizar contacto: ' + str(e)}), 500
    finally:
        conexion.close()

# Endpoint para eliminar contacto
@empresas_bp.route('/contactos/<int:id_contacto>', methods=['DELETE'])
@session_validator(tabla="clientes", accion="delete")
def eliminar_contacto(id_contacto):
    conexion = get_connection()

    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Obtener contacto
            cursor.execute("SELECT * FROM contacto WHERE idContacto = %s", (id_contacto,))
            contacto = cursor.fetchone()

            if not contacto:
                return jsonify({'error': 'Contacto no encontrado'}), 404

            # Eliminar contacto
            cursor.execute("DELETE FROM contacto WHERE idContacto = %s", (id_contacto,))
            conexion.commit()

            # Registrar auditoría
            registrar_auditoria(
                g.user_id,
                'delete',
                'contacto',
                id_contacto,
                valores_anteriores=contacto,
                valores_nuevos=None
            )

            return jsonify({'mensaje': 'Contacto eliminado correctamente'}), 200

    except Exception as e:
        conexion.rollback()
        print("Error al eliminar contacto:", e)
        return jsonify({'error': 'Error al eliminar contacto: ' + str(e)}), 500
    finally:
        conexion.close()

# Endpoint para cambiar estado de empresa
@empresas_bp.route('/empresas/<int:id_cliente>/estado', methods=['PATCH'])
@session_validator(tabla="clientes", accion="update")
def cambiar_estado_empresa(id_cliente):
    data = request.get_json()
    if 'estatus' not in data:
        return jsonify({'error': 'Falta el campo estatus'}), 400

    nuevo_estatus = data['estatus']
    if nuevo_estatus not in ['Activo', 'Inactivo']:
        return jsonify({'error': 'Estatus no válido'}), 400

    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Verificar que sea empresa
            cursor.execute("""
                SELECT estatus FROM clientes 
                WHERE idCliente = %s AND tipoCliente = 'Empresa'
            """, (id_cliente,))
            empresa = cursor.fetchone()

            if not empresa:
                return jsonify({'error': 'Empresa no encontrada'}), 404

            # Actualizar estado
            cursor.execute("""
                UPDATE clientes 
                SET estatus = %s 
                WHERE idCliente = %s
            """, (nuevo_estatus, id_cliente))
            conexion.commit()

            # Registrar auditoría
            registrar_auditoria(
                g.user_id,
                'update',
                'clientes',
                id_cliente,
                valores_anteriores={'estatus': empresa['estatus']},
                valores_nuevos={'estatus': nuevo_estatus}
            )

            return jsonify({'mensaje': 'Estado actualizado correctamente'}), 200

    except Exception as e:
        conexion.rollback()
        print("Error al actualizar estado:", e)
        return jsonify({'error': 'Error al actualizar estado'}), 500
    finally:
        conexion.close()

# Endpoint para exportar empresas a PDF
@empresas_bp.route('/empresas/exportar/pdf', methods=['GET'])
@session_validator(tabla="clientes", accion="read")
def exportar_empresas_pdf():
    try:
        # Parámetros de búsqueda
        search = request.args.get('search', '')
        estatus = request.args.get('estatus', 'all')

        # Consulta de datos
        empresas = obtener_empresas_filtradas(search, estatus)

        # Renderizar HTML con template
        html = render_template(
            'empresa_reporte.html',
            empresas=empresas,
            fecha=datetime.now().strftime('%d/%m/%Y %H:%M')
        )

        # Configuración de PDFKit (ajusta si wkhtmltopdf está en otro path)
        config = pdfkit.configuration(wkhtmltopdf='/usr/local/bin/wkhtmltopdf')

        pdf = pdfkit.from_string(html, False, configuration=config)

        return send_file(
            io.BytesIO(pdf),
            as_attachment=True,
            download_name=f'reporte_empresas_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf',
            mimetype='application/pdf'
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Error generando reporte PDF: ' + str(e)}), 500


# Endpoint para exportar empresas a Excel
@empresas_bp.route('/empresas/exportar/excel', methods=['GET'])
@session_validator(tabla="clientes", accion="read")
def exportar_empresas_excel():
    try:
        # Obtener parámetros de filtrado
        search = request.args.get('search', '')
        estatus = request.args.get('estatus', 'all')

        # Obtener empresas
        empresas = obtener_empresas_filtradas(search, estatus)

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Empresas"

        # Estilos
        header_fill = PatternFill(start_color='3b82f6', end_color='3b82f6', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        alignment = Alignment(horizontal='center', vertical='center')
        border = Border(bottom=Side(style='medium'))

        # Encabezados
        headers = [
            "Nombre Comercial",
            "RFC",
            "Razón Social",
            "Teléfono",
            "Email",
            "Domicilio Fiscal",
            "Fecha Registro",
            "Estado"
        ]

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f'{col_letter}1']
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = border

        # Datos
        for row_num, empresa in enumerate(empresas, 2):
            ws.cell(row=row_num, column=1, value=empresa['nombre'])
            ws.cell(row=row_num, column=2, value=empresa['rfc'] or '')
            ws.cell(row=row_num, column=3, value=empresa['razonSocial'] or '')
            ws.cell(row=row_num, column=4, value=empresa['telefono'] or '')
            ws.cell(row=row_num, column=5, value=empresa['email'])
            ws.cell(row=row_num, column=6, value=empresa['domicilioFiscal'] or '')
            ws.cell(row=row_num, column=7,
                    value=empresa['fechaRegistro'].strftime('%d/%m/%Y') if empresa['fechaRegistro'] else '')
            ws.cell(row=row_num, column=8, value=empresa['estatus'])

        # Ajustar anchos
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 25

        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'reporte_empresas_{datetime.datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Error generando Excel: {str(e)}")
        return jsonify({'error': 'Error generando reporte Excel: ' + str(e)}), 500

# Función auxiliar para obtener empresas con filtros (sin paginación)
def obtener_empresas_filtradas(search, estatus):
    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            query = """
                SELECT c.nombre, c.rfc, c.telefono, c.email, c.fechaRegistro, c.estatus,
                       e.razonSocial, e.domicilioFiscal
                FROM clientes c
                JOIN empresas e ON c.idCliente = e.idCliente
                WHERE c.tipoCliente = 'Empresa'
            """
            params = []

            if search:
                query += " AND (c.nombre LIKE %s OR c.rfc LIKE %s OR c.email LIKE %s OR e.razonSocial LIKE %s)"
                search_term = f"%{search}%"
                params.extend([search_term] * 4)

            if estatus != 'all':
                query += " AND c.estatus = %s"
                params.append(estatus)

            query += " ORDER BY c.fechaRegistro DESC"

            cursor.execute(query, params)
            return cursor.fetchall()

    except Exception as e:
        print(f"Error obteniendo empresas: {str(e)}")
        return []
    finally:
        conexion.close()