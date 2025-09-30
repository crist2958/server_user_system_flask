# client/clientes_personas.py

from flask import request, jsonify, g, Blueprint, send_file
from db_config import get_connection
from utils.session_validator import session_validator
from utils.auditoria import registrar_auditoria
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

import pdfkit
from flask import render_template, request, send_file

personas_bp = Blueprint('personas', __name__)


# Endpoint para registrar cliente persona
@personas_bp.route('/personas', methods=['POST'])
@session_validator(tabla="clientes", accion="create")
def registrar_persona():
    datos = request.json
    campos_requeridos = ['nombre', 'telefono', 'email']

    if not all(campo in datos for campo in campos_requeridos):
        return jsonify({'error': 'Faltan campos requeridos'}), 400

    # Convertir campos vacíos a None para evitar problemas con UNIQUE
    campos_opcionales = ['rfc', 'apellidoP', 'apellidoM', 'direccion']
    for campo in campos_opcionales:
        if campo in datos and datos[campo] == '':
            datos[campo] = None

    conexion = get_connection()
    try:
        with conexion.cursor() as cursor:
            # Insertar en tabla clientes
            cursor.execute("""
                INSERT INTO clientes (
                    tipoCliente, nombre, apellidoP, apellidoM, 
                    rfc, telefono, email, direccion, notas
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                'Persona',
                datos['nombre'],
                datos.get('apellidoP'),
                datos.get('apellidoM'),
                datos.get('rfc'),
                datos['telefono'],
                datos['email'],
                datos.get('direccion'),
                datos.get('notas')
            ))
            id_cliente = cursor.lastrowid

            # Registrar auditoría
            registrar_auditoria(
                g.user_id,
                'create',
                'clientes',
                id_cliente,
                valores_anteriores=None,
                valores_nuevos=datos
            )

            # Insertar direcciones si existen
            direcciones = datos.get('direcciones', [])
            for direccion in direcciones:
                # Convertir campos vacíos en direcciones también
                campos_direccion = ['numero', 'colonia']
                for campo_d in campos_direccion:
                    if campo_d in direccion and direccion[campo_d] == '':
                        direccion[campo_d] = None

                cursor.execute("""
                    INSERT INTO direcciones_envio (
                        idCliente, calle, numero, colonia, 
                        codigoPostal, pais, estado, ciudad
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    id_cliente,
                    direccion['calle'],
                    direccion.get('numero'),
                    direccion.get('colonia'),
                    direccion['codigoPostal'],
                    direccion['pais'],
                    direccion['estado'],
                    direccion['ciudad']
                ))

            conexion.commit()
            return jsonify({
                'mensaje': 'Persona registrada correctamente',
                'idCliente': id_cliente
            }), 201

    except Exception as e:
        conexion.rollback()
        print("Error al registrar persona:", e)

        # Manejar específicamente el error de duplicidad en RFC
        if "Duplicate entry" in str(e) and "rfc" in str(e):
            return jsonify({'error': 'El RFC ya está registrado en el sistema'}), 409

        return jsonify({'error': 'Error al registrar persona: ' + str(e)}), 500
    finally:
        conexion.close()


# Endpoint para actualizar persona
@personas_bp.route('/personas/<int:id_cliente>', methods=['PUT'])
@session_validator(tabla="clientes", accion="update")
def actualizar_persona(id_cliente):
    datos = request.json
    conexion = get_connection()

    # Convertir campos vacíos a None para evitar problemas con UNIQUE
    campos_opcionales = ['rfc', 'apellidoP', 'apellidoM', 'direccion']
    for campo in campos_opcionales:
        if campo in datos and datos[campo] == '':
            datos[campo] = None

    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Verificar que sea una persona
            cursor.execute("SELECT tipoCliente FROM clientes WHERE idCliente = %s", (id_cliente,))
            cliente = cursor.fetchone()

            if not cliente:
                return jsonify({'error': 'Cliente no encontrado'}), 404
            if cliente['tipoCliente'] != 'Persona':
                return jsonify({'error': 'El cliente no es una persona'}), 400

            # Actualizar datos
            campos_actualizables = [
                'nombre', 'apellidoP', 'apellidoM', 'rfc',
                'telefono', 'email', 'direccion', 'notas', 'estatus'
            ]

            valores_anteriores = {}
            valores_nuevos = {}
            set_clause = []
            params = []

            for campo in campos_actualizables:
                if campo in datos:
                    cursor.execute(f"SELECT {campo} FROM clientes WHERE idCliente = %s", (id_cliente,))
                    valor_anterior = cursor.fetchone()[campo]

                    valores_anteriores[campo] = valor_anterior
                    valores_nuevos[campo] = datos[campo]
                    set_clause.append(f"{campo} = %s")
                    params.append(datos[campo])

            if set_clause:
                sql = f"UPDATE clientes SET {', '.join(set_clause)} WHERE idCliente = %s"
                params.append(id_cliente)
                cursor.execute(sql, params)

                # Registrar auditoría
                registrar_auditoria(
                    g.user_id,
                    'update',
                    'clientes',
                    id_cliente,
                    valores_anteriores=valores_anteriores,
                    valores_nuevos=valores_nuevos
                )

            conexion.commit()
            return jsonify({'mensaje': 'Persona actualizada correctamente'}), 200

    except Exception as e:
        conexion.rollback()
        print("Error al actualizar persona:", e)

        # Manejar específicamente el error de duplicidad en RFC
        if "Duplicate entry" in str(e) and "rfc" in str(e):
            return jsonify({'error': 'El RFC ya está registrado en el sistema'}), 409

        return jsonify({'error': 'Error al actualizar persona: ' + str(e)}), 500
    finally:
        conexion.close()

# Endpoint para eliminar persona
@personas_bp.route('/personas/<int:id_cliente>', methods=['DELETE'])
@session_validator(tabla="clientes", accion="delete")
def eliminar_persona(id_cliente):
    conexion = get_connection()

    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Verificar que sea una persona
            cursor.execute("SELECT * FROM clientes WHERE idCliente = %s", (id_cliente,))
            cliente = cursor.fetchone()

            if not cliente:
                return jsonify({'error': 'Cliente no encontrado'}), 404
            if cliente['tipoCliente'] != 'Persona':
                return jsonify({'error': 'El cliente no es una persona'}), 400

            # Eliminar cliente (CASCADE eliminará direcciones)
            cursor.execute("DELETE FROM clientes WHERE idCliente = %s", (id_cliente,))
            conexion.commit()

            # Registrar auditoría
            registrar_auditoria(
                g.user_id,
                'delete',
                'clientes',
                id_cliente,
                valores_anteriores=cliente,
                valores_nuevos=None
            )

            return jsonify({'mensaje': 'Persona eliminada correctamente'}), 200

    except Exception as e:
        conexion.rollback()
        print("Error al eliminar persona:", e)
        return jsonify({'error': 'Error al eliminar persona: ' + str(e)}), 500
    finally:
        conexion.close()


# Endpoint para listar personas
@personas_bp.route('/personas', methods=['GET'])
@session_validator(tabla="clientes", accion="read")
def listar_personas():
    search = request.args.get('search', '')
    estatus = request.args.get('estatus', 'all')

    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            query = """
                SELECT c.*, COUNT(d.idDireccionEnvio) AS num_direcciones
                FROM clientes c
                LEFT JOIN direcciones_envio d ON c.idCliente = d.idCliente
                WHERE c.tipoCliente = 'Persona'
            """
            params = []

            # Filtro de búsqueda
            if search:
                query += " AND (c.nombre LIKE %s OR c.rfc LIKE %s OR c.email LIKE %s)"
                search_term = f"%{search}%"
                params.extend([search_term] * 3)

            # Filtro de estatus
            if estatus != 'all':
                query += " AND c.estatus = %s"
                params.append(estatus)

            query += " GROUP BY c.idCliente ORDER BY c.fechaRegistro DESC"

            cursor.execute(query, params)
            personas = cursor.fetchall()

            # Formatear resultados
            for persona in personas:
                persona['num_direcciones'] = int(persona['num_direcciones'])

            return jsonify(personas), 200

    except Exception as e:
        print("Error al listar personas:", e)
        return jsonify({"error": "Error al listar personas"}), 500
    finally:
        conexion.close()


# Endpoint para obtener detalles de persona
@personas_bp.route('/personas/<int:id_cliente>', methods=['GET'])
@session_validator(tabla="clientes", accion="read")
def obtener_persona(id_cliente):
    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Obtener datos básicos
            cursor.execute("""
                SELECT * FROM clientes 
                WHERE idCliente = %s AND tipoCliente = 'Persona'
            """, (id_cliente,))
            persona = cursor.fetchone()

            if not persona:
                return jsonify({'error': 'Persona no encontrada'}), 404

            # Obtener direcciones
            cursor.execute("""
                SELECT * FROM direcciones_envio 
                WHERE idCliente = %s
            """, (id_cliente,))
            direcciones = cursor.fetchall()

            return jsonify({
                'persona': persona,
                'direcciones': direcciones
            }), 200

    except Exception as e:
        print("Error al obtener persona:", e)
        return jsonify({"error": "Error al obtener persona"}), 500
    finally:
        conexion.close()


# Endpoint para agregar dirección a persona
@personas_bp.route('/personas/<int:id_cliente>/direcciones', methods=['POST'])
@session_validator(tabla="direcciones_envio", accion="create")
def agregar_direccion_persona(id_cliente):
    datos = request.json
    campos_requeridos = ['calle', 'codigoPostal', 'pais', 'estado', 'ciudad']

    if not all(campo in datos for campo in campos_requeridos):
        return jsonify({'error': 'Faltan campos requeridos'}), 400

    conexion = get_connection()
    try:
        with conexion.cursor() as cursor:
            # Verificar que sea persona
            cursor.execute("SELECT tipoCliente FROM clientes WHERE idCliente = %s", (id_cliente,))
            cliente = cursor.fetchone()

            if not cliente or cliente['tipoCliente'] != 'Persona':
                return jsonify({'error': 'Cliente no es una persona'}), 400

            # Insertar dirección
            cursor.execute("""
                INSERT INTO direcciones_envio (
                    idCliente, calle, numero, colonia, 
                    codigoPostal, pais, estado, ciudad
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                id_cliente,
                datos['calle'],
                datos.get('numero'),
                datos.get('colonia'),
                datos['codigoPostal'],
                datos['pais'],
                datos['estado'],
                datos['ciudad']
            ))
            id_direccion = cursor.lastrowid
            conexion.commit()

            # Registrar auditoría
            registrar_auditoria(
                g.user_id,
                'create',
                'direcciones_envio',
                id_direccion,
                valores_anteriores=None,
                valores_nuevos=datos
            )

            return jsonify({
                'mensaje': 'Dirección agregada correctamente',
                'idDireccion': id_direccion
            }), 201

    except Exception as e:
        conexion.rollback()
        print("Error al agregar dirección:", e)
        return jsonify({'error': 'Error al agregar dirección: ' + str(e)}), 500
    finally:
        conexion.close()


# Endpoint para cambiar estado de persona
@personas_bp.route('/personas/<int:id_cliente>/estado', methods=['PATCH'])
@session_validator(tabla="clientes", accion="update")
def cambiar_estado_persona(id_cliente):
    data = request.get_json()
    if 'estatus' not in data:
        return jsonify({'error': 'Falta el campo estatus'}), 400

    nuevo_estatus = data['estatus']
    if nuevo_estatus not in ['Activo', 'Inactivo']:
        return jsonify({'error': 'Estatus no válido'}), 400

    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Verificar que sea persona
            cursor.execute("""
                SELECT estatus FROM clientes 
                WHERE idCliente = %s AND tipoCliente = 'Persona'
            """, (id_cliente,))
            persona = cursor.fetchone()

            if not persona:
                return jsonify({'error': 'Persona no encontrada'}), 404

            # Actualizar estado
            cursor.execute("""
                UPDATE clientes 
                SET estatus = %s 
                WHERE idCliente = %s
            """, (nuevo_estatus, id_cliente))
            conexion.commit()

            # Registrar auditoría
            registrar_auditoria(
                g.user_id,
                'update',
                'clientes',
                id_cliente,
                valores_anteriores={'estatus': persona['estatus']},
                valores_nuevos={'estatus': nuevo_estatus}
            )

            return jsonify({'mensaje': 'Estado actualizado correctamente'}), 200

    except Exception as e:
        conexion.rollback()
        print("Error al actualizar estado:", e)
        return jsonify({'error': 'Error al actualizar estado'}), 500
    finally:
        conexion.close()


# Endpoint para exportar personas a PDF
from flask import render_template, request, send_file, jsonify
from datetime import datetime
import pdfkit
import io

@personas_bp.route('/personas/exportar/pdf', methods=['GET'])
@session_validator(tabla="clientes", accion="read")
def exportar_personas_pdf():
    try:
        # Parámetros de búsqueda
        search = request.args.get('search', '')
        estatus = request.args.get('estatus', 'all')

        # Obtener datos
        personas = obtener_personas_filtradas(search, estatus)

        # Renderizar HTML
        html = render_template(
            'clientes_reporte.html',
            titulo="REPORTE DE PERSONAS",
            fecha=datetime.now().strftime("%d/%m/%Y %H:%M"),
            personas=personas
        )

        # Configuración opcional de wkhtmltopdf (si lo requieres)
        options = {
            'encoding': "UTF-8",
            'enable-local-file-access': None
        }

        # Generar PDF en memoria
        pdf = pdfkit.from_string(html, False, options=options)

        # Enviar archivo
        return send_file(
            io.BytesIO(pdf),
            as_attachment=True,
            download_name=f'reporte_personas_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf',
            mimetype='application/pdf'
        )

    except Exception as e:
        print(f"Error al generar PDF: {e}")
        return jsonify({"error": "Error al generar el reporte PDF"}), 500


# Endpoint para exportar personas a Excel
@personas_bp.route('/personas/exportar/excel', methods=['GET'])
@session_validator(tabla="clientes", accion="read")
def exportar_personas_excel():
    try:
        # Obtener parámetros de filtrado
        search = request.args.get('search', '')
        estatus = request.args.get('estatus', 'all')

        # Obtener personas
        personas = obtener_personas_filtradas(search, estatus)

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Personas"

        # Estilos
        header_fill = PatternFill(start_color='3b82f6', end_color='3b82f6', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        alignment = Alignment(horizontal='center', vertical='center')
        border = Border(bottom=Side(style='medium'))

        # Encabezados
        headers = [
            "Nombre",
            "Apellido Paterno",
            "Apellido Materno",
            "RFC",
            "Teléfono",
            "Email",
            "Dirección",
            "Fecha Registro",
            "Estado"
        ]

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f'{col_letter}1']
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = border

        # Datos
        for row_num, persona in enumerate(personas, 2):
            ws.cell(row=row_num, column=1, value=persona['nombre'])
            ws.cell(row=row_num, column=2, value=persona['apellidoP'] or '')
            ws.cell(row=row_num, column=3, value=persona['apellidoM'] or '')
            ws.cell(row=row_num, column=4, value=persona['rfc'] or '')
            ws.cell(row=row_num, column=5, value=persona['telefono'] or '')
            ws.cell(row=row_num, column=6, value=persona['email'])
            ws.cell(row=row_num, column=7, value=persona['direccion'] or '')
            ws.cell(row=row_num, column=8,
                    value=persona['fechaRegistro'].strftime('%d/%m/%Y') if persona['fechaRegistro'] else '')
            ws.cell(row=row_num, column=9, value=persona['estatus'])

        # Ajustar anchos
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 20

        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'reporte_personas_{datetime.datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Error generando Excel: {str(e)}")
        return jsonify({'error': 'Error generando reporte Excel: ' + str(e)}), 500


# Función auxiliar para obtener personas con filtros (sin paginación)
def obtener_personas_filtradas(search, estatus):
    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            query = """
                SELECT c.*
                FROM clientes c
                WHERE c.tipoCliente = 'Persona'
            """
            params = []

            if search:
                query += " AND (c.nombre LIKE %s OR c.rfc LIKE %s OR c.email LIKE %s)"
                search_term = f"%{search}%"
                params.extend([search_term] * 3)

            if estatus != 'all':
                query += " AND c.estatus = %s"
                params.append(estatus)

            query += " ORDER BY c.fechaRegistro DESC"

            cursor.execute(query, params)
            return cursor.fetchall()

    except Exception as e:
        print(f"Error obteniendo personas: {str(e)}")
        return []
    finally:
        conexion.close()

# Endpoint para eliminar dirección
@personas_bp.route('/direcciones/<int:id_direccion>', methods=['DELETE'])
@session_validator(tabla="direcciones_envio", accion="delete")
def eliminar_direccion(id_direccion):
    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Obtener dirección para auditoría
            cursor.execute("SELECT * FROM direcciones_envio WHERE idDireccionEnvio = %s", (id_direccion,))
            direccion = cursor.fetchone()

            if not direccion:
                return jsonify({'error': 'Dirección no encontrada'}), 404

            # Eliminar dirección
            cursor.execute("DELETE FROM direcciones_envio WHERE idDireccionEnvio = %s", (id_direccion,))
            conexion.commit()

            # Registrar auditoría
            registrar_auditoria(
                g.user_id,
                'delete',
                'direcciones_envio',
                id_direccion,
                valores_anteriores=direccion,
                valores_nuevos=None
            )

            return jsonify({'mensaje': 'Dirección eliminada correctamente'}), 200

    except Exception as e:
        conexion.rollback()
        print("Error al eliminar dirección:", e)
        return jsonify({'error': 'Error al eliminar dirección: ' + str(e)}), 500
    finally:
        conexion.close()


# Endpoint para actualizar dirección
@personas_bp.route('/direcciones/<int:id_direccion>', methods=['PUT'])
@session_validator(tabla="clientes", accion="update")
def actualizar_direccion(id_direccion):
    datos = request.json
    campos_requeridos = ['calle', 'codigoPostal', 'pais', 'estado', 'ciudad']

    if not all(campo in datos for campo in campos_requeridos):
        return jsonify({'error': 'Faltan campos requeridos'}), 400

    conexion = get_connection()
    try:
        with conexion.cursor(dictionary=True) as cursor:
            # Obtener la dirección actual para auditoría
            cursor.execute("SELECT * FROM direcciones_envio WHERE idDireccionEnvio = %s", (id_direccion,))
            direccion_anterior = cursor.fetchone()

            if not direccion_anterior:
                return jsonify({'error': 'Dirección no encontrada'}), 404

            # Convertir campos vacíos a None
            campos_opcionales = ['numero', 'colonia']
            for campo in campos_opcionales:
                if campo in datos and datos[campo] == '':
                    datos[campo] = None

            # Actualizar la dirección
            campos_actualizables = [
                'calle', 'numero', 'colonia', 'codigoPostal',
                'pais', 'estado', 'ciudad'
            ]

            set_clause = []
            params = []
            valores_nuevos = {}
            valores_anteriores = {}

            for campo in campos_actualizables:
                if campo in datos:
                    # Guardar valor anterior para auditoría
                    valores_anteriores[campo] = direccion_anterior[campo]
                    valores_nuevos[campo] = datos[campo]
                    set_clause.append(f"{campo} = %s")
                    params.append(datos[campo])

            if set_clause:
                sql = f"UPDATE direcciones_envio SET {', '.join(set_clause)} WHERE idDireccionEnvio = %s"
                params.append(id_direccion)
                cursor.execute(sql, params)

                # Registrar auditoría
                registrar_auditoria(
                    g.user_id,
                    'update',
                    'direcciones_envio',
                    id_direccion,
                    valores_anteriores=valores_anteriores,
                    valores_nuevos=valores_nuevos
                )

            conexion.commit()
            return jsonify({'mensaje': 'Dirección actualizada correctamente'}), 200

    except Exception as e:
        conexion.rollback()
        print("Error al actualizar dirección:", e)
        return jsonify({'error': 'Error al actualizar dirección: ' + str(e)}), 500
    finally:
        conexion.close()